import pandas 
import openpyxl 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
import os
from pathlib import Path

# mapping JSON key to excel col
mappings_gen_report = {
    'MedicalRec.#':'Medical Record Number (MRN)', #report writer
    'Address1':'Member Residential Address',
    'City':'Member Residential City',
    'Zip':'Member Residential Zip Code',
    'Phone 1':'Member Phone Number',
    'Date of Birth': 'Age Group', #in JSON as DOB, but value in RTF should be binary
    'Service Start':'ECM Benefit Start Date',
    'Status':'Status of Member Engagement',
    'Service End':'ECM Benefit End Date',
    'Case Manager':'ECM Lead Care Manager Name',
    'Authorization Number':'Authorization Number', #Client Details --> Service Orders --> Field: Authorization Number
    'NoteDate':'Date of Outreach Attempt', #Note Date 
    'CIN-CLIENT IINDEX NUMBER': 'Member Client Index Number (CIN)',
    #constant/no extract
    'Provider Type': 'Provider Type',
    'ECM Provider Name':'ECM Provider Name',
    'ECM Provider National Provider Identifier (NPI)':'ECM Provider National Provider Identifier (NPI)',
    'ECM Provider Phone Number': 'ECM Provider Phone Number' 
}  
#--------Populate if data present --------#
conditional_mapping = {
        'NoteType':'Member New Address Indicator', #can have mult note type, filled only if have NT for new address
        'NoteType': 'Member New Phone Number Indicator', #can have mult note type, filled only if have NT for new phone
        'Member Dually Enrolled in Medicare': 'Member Dually Enrolled in Medicare' # can be found in custom fields
}

#not mapped 
conditional_fields = {
        'Member New Address Indicator', #can have mult note type, filled only if have NT for new address
        'Member New Phone Number Indicator', #can have mult note type, filled only if have NT for new phone
        'Member Dually Enrolled in Medicare' # can be found in custom fields
}

#-------- defer to note desc for now--------#
note_desc_fields = [
    'Outreach Attempt Method',
    'Contact Reason',
    'Contact Outcome',
    'Time Spent Performing Outreach',
    'Recommendation for Discontinuation Date',
    'Discontinuation Reason Code'
]

#--------Optional/Not needed--------#  
optional_fields = [
    'Member Preferred Language (Spoken)',
    'Member Preferred Language (Written)',
    'Reauthorization Request',
    'Has Member Been Housed',
    'Housing Type',
    'Date Housed',
]

#---- Serv Auth ----#
serv_auth_fields = [
    'Individuals Experiencing Homelessness',
    'Individuals at Risk for Avoidable Hospital or ED Utilization',
    'Individuals with Serious Mental Health or Substance Use Disorder Needs',
    'Individuals Transitioning from Incarceration',
    'Birth Equity Population of Focus',
    'Individuals Living in the Community and at Risk for LTC Institutionalization',
    'Nursing Facility Residents Transitioning to Community',
    'Enrolled in CCS or CCS WCM with Additional Needs Beyond the CCS Condition',
    'Involved in Child Welfare',
]

#--------manually entered by JM--------#
manual_fields = [
    'ECM Lead Care Manager Phone Number',
    'Member Information RTF Production Date',
    'Member Information RTF Reporting Period',
]

def generate_excel_from_json(json_file_path, output_dir):
    """
    Generate Excel file from JSON data using KP template mapping.

    Args:
        json_file_path: Path to the JSON file to process
        output_dir: Directory to save the output Excel file

    Returns:
        Path to the generated Excel file
    """
    json_file_path = Path(json_file_path)
    output_dir = Path(output_dir)

    # Look for KP template in multiple locations
    template_files = []
    search_paths = [
        Path("."),  # Current directory
        Path.cwd(),  # Working directory
        Path(__file__).parent.parent,  # Project root (backend parent)
    ]

    for search_path in search_paths:
        template_files = list(search_path.glob("*KP*Template*.xlsx"))
        if template_files:
            break

    if template_files:
        print(f"Using template: {template_files[0].name}")
        wb = load_workbook(template_files[0])
        ws = wb.active
    else:
        print("No template found, creating new workbook")
        wb = openpyxl.Workbook()
        ws = wb.active

    # Load JSON data
    with open(json_file_path, "r", encoding='utf-8') as f:
        all_members = json.load(f)

    # Filter to only members with enriched personal_data
    members = [m for m in all_members if m.get('personal_data')]

    print(f"Found {len(all_members)} total records, {len(members)} with personal data to process")

    # Find the header row in the Excel template (or create header row if new workbook)
    if ws.max_row > 1:
        header_row = [cell.value for cell in ws[1]]
    else:
        # Create header row for new workbook with all KP required columns
        header_row = list(mappings_gen_report.values())
        ws.append(header_row)  # Add header row to the worksheet

    # insert data into excel
    for member in members:
        row_data = [None] * len(header_row)
        #fields with other conditions
        for idx, col_name in enumerate(header_row):
            if col_name in note_desc_fields:
                row_data[idx] = 'Refer to NoteDesc in Report Writer'

            elif col_name in conditional_fields:
                row_data[idx] = 'Only populate if data present'

            elif col_name in optional_fields:
                row_data[idx] = 'Optional fields'

            elif col_name in serv_auth_fields:
                row_data[idx] = 'Refer to Service Auth form'

            elif col_name in manual_fields:
                row_data[idx] = 'Entered manually'
        #extractable fields - ENHANCED to handle personal_data nesting
            else:
                # Direct field mapping based on actual JSON structure
                personal_data = member.get('personal_data', {})

                # Map fields from JSON to Excel columns
                if col_name == 'Medical Record Number (MRN)':
                    row_data[idx] = member.get('MedicalRec.#')
                elif col_name == 'Member Residential Address':
                    row_data[idx] = personal_data.get('address_1')
                elif col_name == 'Member Residential City':
                    row_data[idx] = personal_data.get('city')
                elif col_name == 'Member Residential Zip Code':
                    row_data[idx] = personal_data.get('zip')
                elif col_name == 'Member Phone Number':
                    row_data[idx] = personal_data.get('phone_1')
                elif col_name == 'Age Group':
                    row_data[idx] = member.get('DateofBirth')  # Keep as DOB for now
                elif col_name == 'ECM Benefit Start Date':
                    row_data[idx] = personal_data.get('service_start')
                elif col_name == 'Status of Member Engagement':
                    row_data[idx] = member.get('Status')
                elif col_name == 'ECM Benefit End Date':
                    row_data[idx] = personal_data.get('service_end')
                elif col_name == 'ECM Lead Care Manager Name':
                    row_data[idx] = personal_data.get('case_manager')
                elif col_name == 'Authorization Number':
                    row_data[idx] = personal_data.get('referral_number')
                elif col_name == 'Date of Outreach Attempt':
                    row_data[idx] = member.get('NoteDate')
                elif col_name == 'Member Client Index Number (CIN)':
                    row_data[idx] = personal_data.get('med_record')
        ws.append(row_data)

    # format excel sheet (col width)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Save the populated Excel file to output folder
    output_filename = f"ECM_KP_RTF_Template_Populated_{json_file_path.stem}.xlsx"
    output_path = output_dir / output_filename
    wb.save(output_path)
    print("Workbook Saved in:", output_dir.absolute())
    print(f"Filename: {output_filename}")
    print(f"Total records processed: {len(members)}")
    print(f"Columns in template: {len(header_row)}")

    return str(output_path) 
